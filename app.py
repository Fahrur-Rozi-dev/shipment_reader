"""
Flask Web Application for OCR Shipment Extraction.

Provides a premium web UI for uploading PDFs and viewing extraction results
with real-time progress tracking via Server-Sent Events (SSE).
Split-view validation: results on left, PDF page preview on right.
"""
import json
import logging
import os
import re
import time
import uuid
import queue
import threading
import base64
import io
from pathlib import Path

import numpy as np
from PIL import Image

from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    Response,
    send_from_directory,
    send_file,
)

from pdf_splitter import split_pdf
from image_preprocessor import preprocess_image
from ocr_engine import run_ocr
from llm_parser import parse_ocr_text
from spx_parser import parse_spx_text
from page_processor import PageProcessor

# ── App Setup ──────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB max
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.config["PAGES_FOLDER"] = os.path.join(os.path.dirname(__file__), "page_images")

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["PAGES_FOLDER"], exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s │ %(levelname)-8s │ %(name)-20s │ %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── In-memory job storage ──────────────────────────────────────────────────
jobs = {}


def create_job(job_id: str) -> dict:
    """Create a new processing job."""
    job = {
        "id": job_id,
        "status": "pending",
        "progress": 0,
        "total": 0,
        "current_page": 0,
        "message": "Initializing...",
        "result": None,
        "error": None,
        "events": queue.Queue(),
        "page_results": [],  # per-page extraction results
    }
    jobs[job_id] = job
    return job


def send_event(job: dict, event_type: str, data: dict):
    """Push an SSE event to the job's queue."""
    job["events"].put({"event": event_type, "data": json.dumps(data)})


def save_page_image(page_image, job_id: str, page_num: int) -> str:
    """Save a page image as JPEG for preview. Returns the filename."""
    job_dir = os.path.join(app.config["PAGES_FOLDER"], job_id)
    os.makedirs(job_dir, exist_ok=True)

    filename = f"page_{page_num}.jpg"
    filepath = os.path.join(job_dir, filename)

    # Convert to PIL Image if numpy array
    if isinstance(page_image, np.ndarray):
        img = Image.fromarray(page_image)
    else:
        img = page_image

    # Resize for web preview (max 1200px width)
    max_w = 1200
    if img.width > max_w:
        ratio = max_w / img.width
        img = img.resize((max_w, int(img.height * ratio)), Image.LANCZOS)

    img.save(filepath, "JPEG", quality=80)
    return filename


def process_pdf_job(job_id: str, pdf_path: str, courier: str = "jnt"):
    """Background worker: process a PDF and emit progress events."""
    job = jobs[job_id]

    try:
        # ── Step 1: Split PDF ─────────────────────────────────────────
        job["status"] = "processing"
        job["message"] = "Splitting PDF into pages..."
        send_event(job, "progress", {
            "status": "processing",
            "message": "Splitting PDF into pages...",
            "progress": 0, "total": 0,
        })

        pages = split_pdf(pdf_path)
        total = len(pages)
        job["total"] = total

        send_event(job, "progress", {
            "status": "processing",
            "message": f"PDF split into {total} pages. Starting OCR...",
            "progress": 0, "total": total,
        })

        # ── Step 2: Process pages ─────────────────────────────────────
        processor = PageProcessor()
        page_results = []

        for page_num, page_image in enumerate(pages, start=1):
            job["current_page"] = page_num
            job["progress"] = page_num
            job["message"] = f"Processing page {page_num}/{total}..."

            send_event(job, "progress", {
                "status": "processing",
                "message": f"Processing page {page_num}/{total}...",
                "progress": page_num, "total": total, "step": "ocr",
            })

            # Save page image for preview
            save_page_image(page_image, job_id, page_num)

            # Preprocess & OCR
            original_np = np.array(page_image)
            preprocessed = preprocess_image(page_image)
            ocr_result = run_ocr(preprocessed, original_image=original_np)

            if ocr_result.is_mostly_empty:
                llm_result = {"tracking_number": "", "items": []}
            else:
                send_event(job, "progress", {
                    "status": "processing",
                    "message": f"Page {page_num}/{total}: Extracting data...",
                    "progress": page_num, "total": total, "step": "llm",
                })
                llm_result = parse_spx_text(ocr_result.full_text) if courier == "spx" else parse_ocr_text(ocr_result.full_text)

            # Store per-page result
            page_results.append({
                "page": page_num,
                "tracking_number": llm_result.get("tracking_number", ""),
                "items": llm_result.get("items", []),
                "confidence": round(ocr_result.avg_confidence, 1),
                "barcodes": ocr_result.barcodes,
                "is_empty": ocr_result.is_mostly_empty,
            })

            # State update
            processor.process_page(page_num, llm_result, ocr_result)

        # ── Step 3: Finalize ──────────────────────────────────────────
        processor.finalize()

        shipments = processor.get_shipments()
        manual_review = processor.get_manual_review()
        stats = processor.get_stats()

        result = {
            "shipments": shipments,
            "manual_review": manual_review,
            "stats": stats,
            "page_results": page_results,
            "total_pages": total,
            "job_id": job_id,
        }

        job["status"] = "completed"
        job["result"] = result
        job["message"] = "Extraction complete!"

        send_event(job, "complete", {
            "status": "completed",
            "message": "Extraction complete!",
            "result": result,
        })

    except Exception as e:
        logger.error(f"Job {job_id} failed: {e}", exc_info=True)
        job["status"] = "error"
        job["error"] = str(e)
        job["message"] = f"Error: {e}"
        send_event(job, "error", {"status": "error", "message": str(e)})

    finally:
        # Delayed cleanup: delete uploaded PDF + page images after 10 minutes
        # Gives user time to validate results in split-view before cleanup
        def cleanup():
            time.sleep(600)  # 10 minutes
            try:
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                    logger.info(f"Cleaned up uploaded PDF: {pdf_path}")
                page_dir = os.path.join(app.config["PAGES_FOLDER"], job_id)
                if os.path.exists(page_dir):
                    import shutil
                    shutil.rmtree(page_dir)
                    logger.info(f"Cleaned up page images: {page_dir}")
            except Exception as cleanup_err:
                logger.warning(f"Cleanup error: {cleanup_err}")

        threading.Thread(target=cleanup, daemon=True).start()


# ── Routes ─────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload_pdf():
    if "pdf" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["pdf"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400
    if not file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF files are allowed"}), 400

    job_id = str(uuid.uuid4())[:8]
    filename = f"{job_id}_{file.filename}"
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(filepath)

    # Get courier type from form data
    courier = request.form.get('courier', 'jnt').lower()
    if courier not in ('jnt', 'spx'):
        courier = 'jnt'

    job = create_job(job_id)
    job['courier'] = courier
    thread = threading.Thread(
        target=process_pdf_job, args=(job_id, filepath, courier), daemon=True
    )
    thread.start()
    return jsonify({"job_id": job_id, "courier": courier, "message": "Processing started"})


@app.route("/progress/<job_id>")
def progress_stream(job_id):
    if job_id not in jobs:
        return jsonify({"error": "Job not found"}), 404

    def generate():
        job = jobs[job_id]
        while True:
            try:
                event = job["events"].get(timeout=30)
                yield f"event: {event['event']}\ndata: {event['data']}\n\n"
                if event["event"] in ("complete", "error"):
                    break
            except queue.Empty:
                yield f"event: ping\ndata: {{}}\n\n"

    return Response(
        generate(), mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/result/<job_id>")
def get_result(job_id):
    if job_id not in jobs:
        return jsonify({"error": "Job not found"}), 404
    job = jobs[job_id]
    return jsonify({
        "status": job["status"],
        "result": job["result"],
        "error": job["error"],
        "message": job["message"],
        "progress": job.get("progress", 0),
        "total": job.get("total", 0),
    })


@app.route("/page-image/<job_id>/<int:page_num>")
def page_image(job_id, page_num):
    """Serve a page image for the split-view preview."""
    job_dir = os.path.join(app.config["PAGES_FOLDER"], job_id)
    filename = f"page_{page_num}.jpg"
    filepath = os.path.join(job_dir, filename)

    if not os.path.exists(filepath):
        return jsonify({"error": "Page image not found"}), 404

    return send_file(filepath, mimetype="image/jpeg")


@app.route("/download/<job_id>")
def download_result(job_id):
    if job_id not in jobs:
        return jsonify({"error": "Job not found"}), 404
    job = jobs[job_id]
    if job["status"] != "completed" or not job["result"]:
        return jsonify({"error": "Result not ready"}), 400
    return Response(
        json.dumps(job["result"], indent=2, ensure_ascii=False),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename=shipments_{job_id}.json"},
    )


# ── SKU Transform ──────────────────────────────────────────────────────────
# DSM-8cm-100pcs → KD8100  (DSM → KD, 8 = cm size, 100 = pcs count)
# Non-DSM SKUs remain unchanged
_DSM_PATTERN = re.compile(
    r"^DSM[- ](\d+)\s*c?e?m[- ](\d+)\s*(?:pcs|pes|pce|pss)",
    re.IGNORECASE
)

def transform_sku(raw_sku: str) -> str:
    """Transform DSM-style SKU to KD format. Non-DSM SKUs pass through."""
    m = _DSM_PATTERN.match(raw_sku.strip())
    if m:
        cm_size = m.group(1)
        pcs_count = m.group(2)
        return f"KD{cm_size}{pcs_count}"
    return raw_sku


@app.route("/download-excel/<job_id>")
def download_excel(job_id):
    """Download extraction results as an Excel spreadsheet."""
    if job_id not in jobs:
        return jsonify({"error": "Job not found"}), 404
    job = jobs[job_id]
    if job["status"] != "completed" or not job["result"]:
        return jsonify({"error": "Result not ready"}), 400

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"

    # ── Header style ──
    header_font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7C5CF0", end_color="5A3DD6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    # ── Write headers ──
    headers = ["No Resi", "SKU", "Qty"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ── Write data ──
    row = 2
    shipments = job["result"].get("shipments", [])
    for shipment in shipments:
        tracking = shipment.get("tracking_number", "")
        items = shipment.get("items", [])

        if not items:
            # Shipment without items
            ws.cell(row=row, column=1, value=tracking)
            ws.cell(row=row, column=2, value="-")
            ws.cell(row=row, column=3, value=0)
            row += 1
        else:
            for item in items:
                ws.cell(row=row, column=1, value=tracking)
                raw_sku = item.get("variant", item.get("name", "-"))
                ws.cell(row=row, column=2, value=transform_sku(raw_sku))
                ws.cell(row=row, column=3, value=item.get("quantity", 1))
                row += 1

    # ── Column widths ──
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 8

    # ── Save to memory buffer ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"shipments_{job_id}.xlsx",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 7860))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(debug=debug, host="0.0.0.0", port=port, threaded=True)
